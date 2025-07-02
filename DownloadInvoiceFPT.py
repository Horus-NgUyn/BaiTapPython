# import
import os
import time
import shutil
import xml.etree.ElementTree as ET
import pandas as pd
from urllib.parse import urlparse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import load_workbook, Workbook

# Bước 1: Cấu hình thư mục và trình duyệt
def setup_driver(folder_download_dir):
    # Cấu hình thư muc
    os.makedirs(folder_download_dir, exist_ok=True)
    options = Options()
    options.add_experimental_option("prefs", {
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "download.default_directory": folder_download_dir,
        "plugins.always_open_pdf_externally": True,
        "profile.default_content_settings.popups": 0,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1
    })
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    service = Service()
    driver = webdriver.Chrome(service=service, options=options)
    driver.maximize_window()
    return driver, WebDriverWait(driver, 10)

# Bước 2: Tra cứu hóa đơn
def tra_cuu_hoa_don(driver, wait,ma_so_thue, ma_tra_cuu, url ):
    try:
        driver.get(url)
        # FPT
        if "https://tracuuhoadon.fpt.com.vn/search.html" in url:
            ma_so_thue = str(ma_so_thue).strip().replace("'", "")
            ma_tra_cuu = str(ma_tra_cuu).strip()
            # Nhập mã số thuế
            mst_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='MST bên bán']")))
            driver.execute_script("arguments[0].scrollIntoView(true);", mst_input)
            mst_input.clear()
            mst_input.send_keys(ma_so_thue)
            # Mã tra cứu
            mtc_input = driver.find_element(By.XPATH, "//input[@placeholder='Mã tra cứu hóa đơn']")
            driver.execute_script("arguments[0].scrollIntoView(true);", mtc_input)
            mtc_input.clear()
            mtc_input.send_keys(ma_tra_cuu)
            # Button tra cứu
            print("Đang nhấn nút tra cứu...")
            btn_fpt_search = driver.find_element(By.XPATH,"//button[contains(@class, 'webix_button') and contains(text(), 'Tra cứu')]")
            time.sleep(0.5)
            btn_fpt_search.click()
        # Misa
        elif "https://www.meinvoice.vn/tra-cuu/" in url:
            # Nhập mã hóa đơn
            mtc_input_misa = wait.until(EC.presence_of_element_located((By.NAME, "txtCode")))
            # Đề phòng bị ghi đè
            driver.execute_script("""
                   const header = document.querySelector('.top-header');
                   if (header) header.style.display = 'none';
                   arguments[0].scrollIntoView({block: 'center'});
               """, mtc_input_misa)
            time.sleep(0.5)
            driver.execute_script("arguments[0].value = '';", mtc_input_misa)
            mtc_input_misa.clear()
            mtc_input_misa.send_keys(ma_tra_cuu)
            # Button tra cứu
            print("Đang nhấn nút tra cứu...")
            btn_search_misa = wait.until(EC.element_to_be_clickable((By.ID, "btnSearchInvoice")))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn_search_misa)
            time.sleep(0.5)
            btn_search_misa.click()
        # EHoaDon
        elif "https://van.ehoadon.vn/TCHD?MTC=" in url:
            # Nhập mã hóa đơn
            mts_ehoadon = wait.until(EC.presence_of_element_located((By.ID, "txtInvoiceCode")))
            driver.execute_script("arguments[0].scrollIntoView(true);", mts_ehoadon)
            mts_ehoadon.clear()
            mts_ehoadon.send_keys(ma_tra_cuu)
            # Button tra cứu
            print("Đang nhấn nút tra cứu...")
            btn_ehoadon_search = driver.find_element(By.CLASS_NAME, "btnSearch")
            btn_ehoadon_search.click()
        print("Đang chờ kết quả hóa đơn hiển thị")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
    except TimeoutException:
        print("Không tìm thấy hóa đơn")

# Bước 3: Tải hóa đơn XML
def tai_file_xml(driver,wait, folder_download_dir, url, ma_tra_cuu):
    try:
        # FPT
        if "https://tracuuhoadon.fpt.com.vn/search.html" in url:
            button_fpt = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//button[span[contains(@class, 'mdi-xml')] and contains(text(), 'Tải XML')]"))
            )
            button_fpt.click()
            print("Đã nhấn nút tải file XML từ FPT")
            time.sleep(2)
        # Misa
        elif "https://www.meinvoice.vn/tra-cuu/" in url:
            btn_misa = wait.until(
                EC.element_to_be_clickable((By.CLASS_NAME, "download"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", btn_misa)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", btn_misa)
            print("Đã nhấn mở menu")
            btn_down_xml_meinvoice = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "txt-download-xml")))
            driver.execute_script("arguments[0].click();", btn_down_xml_meinvoice)
            print("Đã chọn tải hóa đơn file XML")
            time.sleep(2)
        # EHoaDon - Improved handling
        elif "https://van.ehoadon.vn/TCHD?MTC=" in url:
            print("Di chuyển vào iframe của EHoaDon")
            iframe_found = False
            
            # Wait for page to load completely
            time.sleep(3)
            
            # Multiple iframe detection strategies
            iframe_strategies = [
                (By.ID, "frameViewInvoice"),
                (By.NAME, "frameViewInvoice"),
                (By.TAG_NAME, "iframe"),
                (By.CSS_SELECTOR, "iframe[src*='invoice']"),
                (By.CSS_SELECTOR, "iframe[id*='frame']"),
                (By.CSS_SELECTOR, "iframe[name*='frame']")
            ]
            
            for strategy in iframe_strategies:
                try:
                    print(f"Thử chiến lược iframe: {strategy}")
                    # First check if iframe exists
                    iframes = driver.find_elements(*strategy)
                    if iframes:
                        print(f"Tìm thấy {len(iframes)} iframe(s)")
                        # Try to switch to the first one
                        wait.until(EC.frame_to_be_available_and_switch_to_it(strategy))
                        iframe_found = True
                        print("Đã vào iframe thành công")
                        break
                except Exception as e:
                    print(f"Không thể vào iframe với strategy {strategy}: {e}")
                    continue
            
            if not iframe_found:
                print("Không tìm thấy iframe nào, thử trực tiếp trên trang chính")
                # Try without iframe
                pass
            
            # Multiple download button strategies
            download_success = False
            download_strategies = [
                (By.ID, "btnDownload"),
                (By.CLASS_NAME, "btnDownload"),
                (By.XPATH, "//button[contains(@id, 'Download')]"),
                (By.XPATH, "//a[contains(@id, 'Download')]"),
                (By.XPATH, "//button[contains(text(), 'Tải')]"),
                (By.XPATH, "//a[contains(text(), 'Tải')]"),
                (By.CSS_SELECTOR, "[id*='download']"),
                (By.CSS_SELECTOR, "[class*='download']")
            ]
            
            for strategy in download_strategies:
                try:
                    print(f"Thử tìm nút download với: {strategy}")
                    btn_download = wait.until(EC.presence_of_element_located(strategy))
                    print("Tìm thấy nút download, đang hover...")
                    
                    # Scroll to element
                    driver.execute_script("arguments[0].scrollIntoView(true);", btn_download)
                    time.sleep(1)
                    
                    # Hover over download button
                    ActionChains(driver).move_to_element(btn_download).perform()
                    time.sleep(1)
                    print("Đã hover nút Download")
                    
                    # Try multiple ways to show dropdown menu
                    dropdown_strategies = [
                        "document.querySelector('#divDownloads .dropdown-menu').style.display='block';",
                        "document.querySelector('.dropdown-menu').style.display='block';",
                        "document.querySelector('[class*=\"dropdown\"]').style.display='block';",
                        "arguments[0].click();"  # Direct click on download button
                    ]
                    
                    for dropdown_script in dropdown_strategies:
                        try:
                            if dropdown_script == "arguments[0].click();":
                                driver.execute_script(dropdown_script, btn_download)
                            else:
                                driver.execute_script(dropdown_script)
                            time.sleep(1)
                            
                            # Try to find XML download link
                            xml_link_strategies = [
                                (By.ID, "LinkDownXML"),
                                (By.CLASS_NAME, "LinkDownXML"),
                                (By.XPATH, "//a[contains(@id, 'XML')]"),
                                (By.XPATH, "//a[contains(text(), 'XML')]"),
                                (By.XPATH, "//button[contains(text(), 'XML')]"),
                                (By.CSS_SELECTOR, "[id*='XML']"),
                                (By.CSS_SELECTOR, "[href*='xml']")
                            ]
                            
                            for xml_strategy in xml_link_strategies:
                                try:
                                    xml_link = driver.find_element(*xml_strategy)
                                    if xml_link.is_displayed() or xml_link.is_enabled():
                                        xml_link.click()
                                        print("Đã chọn tải hóa đơn dạng XML")
                                        download_success = True
                                        time.sleep(2)
                                        break
                                except:
                                    continue
                            
                            if download_success:
                                break
                                
                        except Exception as e:
                            print(f"Lỗi với dropdown strategy: {e}")
                            continue
                    
                    if download_success:
                        break
                        
                except Exception as e:
                    print(f"Lỗi với download strategy {strategy}: {e}")
                    continue
            
            if not download_success:
                print("Không tìm thấy nút tải XML sau khi thử tất cả strategies")
                # Switch back to default content before returning
                try:
                    driver.switch_to.default_content()
                except:
                    pass
                return None
            
            # Switch back to default content
            try:
                driver.switch_to.default_content()
                print("Đã thoát khỏi iframe về trang chính")
            except:
                pass
                
    except TimeoutException:
        print("Không tìm thấy nút tải XML hoặc mã tra cứu không đúng.")
        try:
            driver.switch_to.default_content()
        except:
            pass
        return None
    except Exception as e:
        print("Lỗi khi click nút tải:", e)
        try:
            driver.switch_to.default_content()
        except:
            pass
        return None
        
    # Tạo thư mục riêng cho từng file hóa đơn
    folder = urlparse(url).netloc.replace("www.", "")
    domain_folder = os.path.join(folder_download_dir, folder)
    os.makedirs(domain_folder, exist_ok=True)
    
    # Tìm file XML vừa tải với retry mechanism
    print("Đang chờ file XML được tải...")
    for attempt in range(15):  # Tăng thời gian chờ
        files = os.listdir(folder_download_dir)
        xml_files = [f for f in files if f.endswith(".xml") and not f.startswith("~")]
        
        if xml_files:
            # Get the newest XML file
            xml_files.sort(key=lambda x: os.path.getctime(os.path.join(folder_download_dir, x)), reverse=True)
            newest_xml = xml_files[0]
            src = os.path.join(folder_download_dir, newest_xml)
            dest = os.path.join(domain_folder, f"{ma_tra_cuu}.xml")
            
            try:
                shutil.move(src, dest)
                print(f"Đã lưu file XML: {dest}")
                return dest
            except Exception as e:
                print(f"Lỗi khi di chuyển file: {e}")
                # Try copy instead of move
                try:
                    shutil.copy2(src, dest)
                    os.remove(src)
                    print(f"Đã lưu file XML (copy): {dest}")
                    return dest
                except:
                    pass
        
        time.sleep(1)
        print(f"Thử lại lần {attempt + 1}/15...")
    
    print("Không tìm thấy file XML vừa tải")
    return None

# Bước 4:  Đọc dữ liệu từ file XML
def read_invoice_xml(xml_file_path):
    try:
        # B1: Phân tích XML
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        # B2: Ưu tiên tìm node HDon/DLHDon
        hdon_node = root.find(".//HDon")
        invoice_node = hdon_node.find("DLHDon") if hdon_node is not None else None
        # B3: Nếu không tìm thấy, thử lần lượt các node phổ biến khác
        if invoice_node is None:
            for tag in [".//DLHDon", ".//TDiep", ".//Invoice"]:
                node = root.find(tag)
                if node is not None:
                    invoice_node = node
                    break
            else:
                print(f"Không xác định được node dữ liệu chính trong file: {os.path.basename(xml_file_path)}")
                return None
        # B4: Hàm tìm nhanh theo path kiểu "NDHDon/NBan/Ten"
        def find(path):
            current = invoice_node
            for part in path.split("/"):
                if current is not None:
                    current = current.find(part)
                else:
                    return None
            return current.text if current is not None else None
        # B5: Lấy thông tin số tài khoản bán (có thể nằm trong TTKhac)
        stk_ban = find("NDHDon/NBan/STKNHang")
        if not stk_ban:
            for thongtin in invoice_node.findall(".//NBan/TTKhac/TTin"):
                if thongtin.findtext("TTruong") == "SellerBankAccount":
                    stk_ban = thongtin.findtext("DLieu")
                    break
        # B6: Trả về thông tin cần thiết từ XML
        return {
            'Số hóa đơn': find("TTChung/SHDon"),
            'Đơn vị bán hàng': find("NDHDon/NBan/Ten"),
            'Mã số thuế bán': find("NDHDon/NBan/MST"),
            'Địa chỉ bán': find("NDHDon/NBan/DChi"),
            'Số tài khoản bán': stk_ban,
            'Họ tên người mua hàng': find("NDHDon/NMua/Ten"),
            'Địa chỉ mua': find("NDHDon/NMua/DChi"),
            'Mã số thuế mua': find("NDHDon/NMua/MST"),
        }
    except Exception as error:
        print(f"Lỗi khi đọc file XML {os.path.basename(xml_file_path)}: {error}")
        return None

# Bước 5: Ghi ra thông tin đã đọc vào excel
def append_to_excel(filepath, row_data):
    if not os.path.isfile(filepath):
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Invoices")
        else:
            ws.title = "Invoices"
        ws.append([
            "STT", "Mã số thuế", "Mã tra cứu", "URL",
            "Số hóa đơn", "Đơn vị bán hàng", "Mã số thuế bán", "Địa chỉ bán", "Số tài khoản bán",
            "Họ tên người mua hàng", "Địa chỉ mua", "Mã số thuế mua"
        ])
        wb.save(filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Invoices")
    ws.append(row_data)
    wb.save(filepath)

# Hàm main chính
def main():
    input_file = "input.xlsx"
    output_file = "output_information_invoice.xlsx"
    folder_download_dir = os.path.join(os.getcwd(),"Download_Invoice_Folder")
    driver, wait = setup_driver(folder_download_dir)
    df_invoice = pd.read_excel(input_file, dtype=str)
    for stt, (index, row) in enumerate(df_invoice.iterrows(), 1):
        ma_so_thue = row.get("Mã số thuế", "")
        ma_so_thue = str(ma_so_thue).strip() if ma_so_thue and ma_so_thue != 'nan' else ""
        ma_tra_cuu = str(row.get("Mã tra cứu", "") or "").strip()
        url = str(row.get("URL", "") or "").strip()
        if not url or not ma_tra_cuu:
            continue
        print(f"\n Đang tra cứu mã: {ma_tra_cuu} | Trang web: {url}")
        tra_cuu_hoa_don(driver, wait,ma_so_thue, ma_tra_cuu, url)
        xml_path = tai_file_xml(driver,wait, folder_download_dir, url, ma_tra_cuu)
        if xml_path:
            parsed = read_invoice_xml(xml_path)
            if parsed:
                row_data = [stt, ma_so_thue, ma_tra_cuu, url] + list(parsed.values()) + [""]
            else:
                row_data = [stt, ma_so_thue, ma_tra_cuu, url] + [""] * 9 + [os.path.basename(xml_path), "Không đọc được XML"]
        else:
            row_data = [stt, ma_so_thue, ma_tra_cuu, url] + [""] * 9 + ["", "Fail"]
        append_to_excel(output_file, row_data)
    driver.quit()
    print(f"======Đã đọc hoàn thành và lưu lại ở {output_file}======")

if __name__ == "__main__":
    main()